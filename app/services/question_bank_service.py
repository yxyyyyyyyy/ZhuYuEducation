from __future__ import annotations

from collections import defaultdict
import csv
from datetime import datetime
import io
import json
import re
import uuid

from sqlalchemy import delete, or_, select

from app.core.database import (
    ClassroomORM,
    KnowledgeNodeORM,
    PracticeRecordORM,
    QuestionBankORM,
    StudentMasteryORM,
    StudentProfileORM,
    UserORM,
)
from app.domain.models import (
    CsvImportResponse,
    ImportFailureItem,
    KNOWLEDGE_TIERS,
    PracticeCoachCard,
    PracticeCoachRequest,
    PracticeAnalyticsSummary,
    PracticeAnalyticsTopic,
    PracticeReviewResolveRequest,
    PracticeReviewView,
    PracticeSubmissionRequest,
    PracticeSubmissionResponse,
    Question,
    QuestionBankImportRequest,
    QuestionBankImportItem,
    QuestionBankItemView,
    QuestionGenerateRequest,
    QuestionGenerateResponse,
    QuestionOption,
    QuestionReviewRequest,
    QuestionReviewResponse,
    QuestionType,
    ScorePoint,
    ScorePointResult,
    SimilarQuestionView,
    Topic,
    WorkedStep,
)
from app.repositories.knowledge_repository import KnowledgeRepository
from app.repositories.sql_repository import sql_repository


class QuestionBankService:
    def __init__(self, repository: KnowledgeRepository, llm_service=None) -> None:
        self.repository = repository
        self.llm_service = llm_service
        self._knowledge_tier_set = set(KNOWLEDGE_TIERS)
        self._excel_template_headers = [
            "题目",
            "答案",
            "二级知识点",
            "题型",
            "解析",
            "难度级别",
            "选项",
            "空数",
            "得分点",
            "标签",
        ]

    def import_questions(self, request: QuestionBankImportRequest) -> list[QuestionBankItemView]:
        imported = []
        with sql_repository.session() as session:
            for item in request.questions:
                row = session.execute(
                    select(QuestionBankORM).where(QuestionBankORM.external_id == item.id)
                ).scalars().first()
                if row is None:
                    row = QuestionBankORM(external_id=item.id)
                    session.add(row)
                knowledge_l1_id, knowledge_l2_id = self._validate_knowledge_binding(
                    session,
                    item.knowledge_l1_id,
                    item.knowledge_l2_id,
                )
                row.knowledge_l1_id = knowledge_l1_id
                row.knowledge_l2_id = knowledge_l2_id
                row.topic_id = knowledge_l2_id
                row.stem = item.stem
                row.difficulty_level = int(item.difficulty_level)
                row.difficulty = (
                    float(item.difficulty)
                    if item.difficulty is not None
                    else self._difficulty_level_to_float(item.difficulty_level)
                )
                row.answer = item.answer
                row.explanation = item.explanation
                row.knowledge_tiers = self._normalize_knowledge_tiers(item.knowledge_tiers)
                row.question_type = item.question_type.value
                row.options = [self._dump_model(option) for option in item.options]
                row.blank_count = item.blank_count
                row.score_points = [self._dump_model(point) for point in item.score_points]
                row.tags = item.tags
                session.flush()
                imported.append(self._view(session, row))
        return imported

    def list_questions(self) -> list[QuestionBankItemView]:
        with sql_repository.session() as session:
            rows = session.execute(select(QuestionBankORM).order_by(QuestionBankORM.created_at.desc())).scalars().all()
            return self._views(session, rows)

    def list_questions_by_topic(self, topic_id: str) -> list[Question]:
        with sql_repository.session() as session:
            rows = session.execute(
                select(QuestionBankORM).where(
                    (QuestionBankORM.knowledge_l2_id == topic_id) | (QuestionBankORM.topic_id == topic_id)
                )
            ).scalars().all()
            return [self._question_from_row(row) for row in rows]

    def get_question(self, external_id: str) -> Question | None:
        with sql_repository.session() as session:
            row = session.execute(
                select(QuestionBankORM).where(QuestionBankORM.external_id == external_id)
            ).scalars().first()
            if not row:
                return None
            return self._question_from_row(row)

    def submit_practice(
        self,
        student_profile_id: int,
        recommended_band: str,
        request: PracticeSubmissionRequest,
    ) -> PracticeSubmissionResponse:
        question = self.get_question(request.question_id)
        if not question:
            raise ValueError("question not found")
        evaluation = self._evaluate_answer(question, request.student_answer, request.blank_answers)
        is_correct = evaluation["is_correct"]
        score = evaluation["score"]
        review_status = evaluation.get("review_status", "graded")
        mastery_applied = review_status == "graded"
        mastery_delta = self._mastery_delta_for_score(score) if mastery_applied else 0.0
        stored_answer = "，".join(request.blank_answers) if request.blank_answers is not None else request.student_answer

        record_id = None
        with sql_repository.session() as session:
            record = PracticeRecordORM(
                student_profile_id=student_profile_id,
                question_external_id=question.id,
                topic_id=question.topic_id,
                recommended_band=recommended_band,
                student_answer=stored_answer,
                is_correct=1 if is_correct else 0,
                score=score,
                earned_points=evaluation["earned_points"],
                total_points=evaluation["total_points"],
                evaluation_method=evaluation["method"],
                feedback=evaluation["feedback"],
                evaluation_status="pending_review" if review_status == "pending_review" else "graded",
                review_reason=evaluation.get("review_reason", ""),
                mastery_applied=1 if mastery_applied else 0,
                duration_seconds=request.duration_seconds,
            )
            session.add(record)
            session.flush()
            record_id = record.id
            if mastery_applied:
                self._apply_mastery_delta(session, student_profile_id, question.topic_id, is_correct, score)

        return PracticeSubmissionResponse(
            question_id=question.id,
            is_correct=is_correct,
            correct_answer=question.answer,
            explanation=question.explanation,
            mastery_delta=round(mastery_delta, 2),
            answer_type=question.question_type,
            earned_points=evaluation["earned_points"],
            total_points=evaluation["total_points"],
            score=round(score, 2),
            score_label=evaluation["score_label"],
            evaluation_method=evaluation["method"],
            feedback=evaluation["feedback"],
            breakdown=evaluation["breakdown"],
            review_status=review_status,
            review_record_id=record_id if review_status == "pending_review" else None,
            review_reason=evaluation.get("review_reason", ""),
        )

    def analytics_for_students(self, student_ids: list[int]) -> PracticeAnalyticsSummary:
        with sql_repository.session() as session:
            rows = session.execute(
                select(PracticeRecordORM).where(
                    PracticeRecordORM.student_profile_id.in_(student_ids),
                    PracticeRecordORM.evaluation_status != "pending_review",
                )
            ).scalars().all()

        total = len(rows)
        correct = sum(1 for row in rows if row.is_correct)
        grouped = defaultdict(list)
        for row in rows:
            grouped[row.topic_id].append(row)

        topics = []
        for topic_id, items in grouped.items():
            accuracy = sum(1 for item in items if item.is_correct) / len(items)
            avg_duration = sum(item.duration_seconds for item in items) / len(items) if items else 0
            topics.append(
                PracticeAnalyticsTopic(
                    topic_id=topic_id,
                    attempt_count=len(items),
                    accuracy=round(accuracy, 2),
                    avg_duration_seconds=round(avg_duration, 1),
                )
            )
        topics.sort(key=lambda item: item.attempt_count, reverse=True)
        return PracticeAnalyticsSummary(
            total_attempts=total,
            correct_attempts=correct,
            accuracy=round((correct / total), 2) if total else 0.0,
            topics=topics,
        )

    def list_practice_reviews(self, teacher_user_id: int, status: str = "pending") -> list[PracticeReviewView]:
        with sql_repository.session() as session:
            students = session.execute(
                select(StudentProfileORM).where(StudentProfileORM.user_id == teacher_user_id)
            ).scalars().all()
            student_map = {student.id: student.name for student in students}
            if not student_map:
                return []

            query = select(PracticeRecordORM).where(PracticeRecordORM.student_profile_id.in_(list(student_map.keys())))
            if status == "reviewed":
                query = query.where(PracticeRecordORM.evaluation_status == "reviewed")
            elif status == "all":
                query = query.where(PracticeRecordORM.evaluation_status.in_(["pending_review", "reviewed"]))
            else:
                query = query.where(PracticeRecordORM.evaluation_status == "pending_review")
            records = session.execute(query.order_by(PracticeRecordORM.created_at.desc())).scalars().all()
            question_ids = [record.question_external_id for record in records]
            questions = {
                row.external_id: row
                for row in session.execute(
                    select(QuestionBankORM).where(QuestionBankORM.external_id.in_(question_ids))
                ).scalars().all()
            } if question_ids else {}
            return [
                self._practice_review_view(record, student_map.get(record.student_profile_id, "学生"), questions.get(record.question_external_id))
                for record in records
            ]

    def resolve_practice_review(
        self,
        record_id: int,
        teacher_user_id: int,
        request: PracticeReviewResolveRequest,
    ) -> PracticeReviewView:
        with sql_repository.session() as session:
            record = session.execute(
                select(PracticeRecordORM).where(PracticeRecordORM.id == record_id)
            ).scalars().first()
            if not record:
                raise ValueError("practice review not found")
            student = session.execute(
                select(StudentProfileORM).where(
                    StudentProfileORM.id == record.student_profile_id,
                    StudentProfileORM.user_id == teacher_user_id,
                )
            ).scalars().first()
            if not student:
                raise ValueError("practice review not found")

            score = max(0.0, min(1.0, float(request.score)))
            total_points = record.total_points or 1.0
            record.score = score
            record.earned_points = round(score * total_points, 2)
            record.is_correct = 1 if request.is_correct else 0
            record.feedback = request.feedback or ("教师复核判定为正确。" if request.is_correct else "教师复核判定为需订正。")
            record.evaluation_method = "teacher_review"
            record.evaluation_status = "reviewed"
            record.review_reason = ""
            record.reviewed_at = datetime.utcnow()
            record.reviewed_by_user_id = teacher_user_id
            if not record.mastery_applied:
                self._apply_mastery_delta(session, record.student_profile_id, record.topic_id, bool(record.is_correct), score)
                record.mastery_applied = 1

            question = session.execute(
                select(QuestionBankORM).where(QuestionBankORM.external_id == record.question_external_id)
            ).scalars().first()
            session.flush()
            return self._practice_review_view(record, student.name, question)

    def _mastery_delta_for_score(self, score: float) -> float:
        return 0.05 if score >= 0.85 else (0.01 if score >= 0.55 else -0.03)

    def _apply_mastery_delta(
        self,
        session,
        student_profile_id: int,
        topic_id: str,
        is_correct: bool,
        score: float,
    ) -> None:
        mastery_delta = self._mastery_delta_for_score(score)
        mastery_row = session.execute(
            select(StudentMasteryORM).where(
                StudentMasteryORM.student_profile_id == student_profile_id,
                StudentMasteryORM.topic_id == topic_id,
            )
        ).scalars().first()
        if mastery_row:
            mastery_row.practice_count += 1
            if is_correct:
                mastery_row.correct_count += 1
            mastery_row.mastery = max(0.0, min(1.0, mastery_row.mastery + mastery_delta))

    def _practice_review_view(
        self,
        record: PracticeRecordORM,
        student_name: str,
        question: QuestionBankORM | None,
    ) -> PracticeReviewView:
        return PracticeReviewView(
            record_id=record.id,
            student_profile_id=record.student_profile_id,
            student_name=student_name,
            question_id=record.question_external_id,
            topic_id=record.topic_id,
            question_stem=question.stem if question else "题目已删除",
            correct_answer=question.answer if question else "",
            explanation=question.explanation if question else "",
            student_answer=record.student_answer,
            score=record.score or 0.0,
            is_correct=bool(record.is_correct),
            evaluation_method=record.evaluation_method or "pending_teacher_review",
            feedback=record.feedback or "",
            evaluation_status=record.evaluation_status or "pending_review",
            review_reason=record.review_reason or "",
            created_at=record.created_at,
            reviewed_at=record.reviewed_at,
        )

    def build_coach_card(self, request: PracticeCoachRequest) -> PracticeCoachCard:
        question = self.get_question(request.question_id)
        if not question:
            raise ValueError("question not found")
        topic = self._safe_topic(question.topic_id)
        mismatch = bool(
            request.student_answer
            and request.student_answer.strip().lower() != question.answer.strip().lower()
        )
        step_cards = self._build_step_cards(question, topic.name)
        if mismatch:
            step_cards.insert(
                1,
                WorkedStep(
                    title="先纠正当前误区",
                    content=(
                        f"你当前填写的是“{request.student_answer}”，和正确答案“{question.answer}”不一致。"
                        " 建议回到定义层面重新判断变量关系或公式含义。"
                    ),
                ),
            )

        misconception_alerts = list(dict.fromkeys((topic.common_mistakes or []) + (question.tags or [])))[:4]
        next_drills = [
            f"围绕 {topic.name} 再刷 2 道同类基础题，重点盯住同一类关键词。",
            f"把这题总结成一句规则：{question.explanation[:36]}..." if question.explanation else f"把 {topic.name} 的定义写成自己的话。",
            f"下一轮优先练习与 {topic.name} 同主题、难度接近 {question.difficulty:.2f} 的题。",
        ]
        return PracticeCoachCard(
            question_id=question.id,
            topic_id=question.topic_id,
            topic_name=topic.name,
            strategy_summary=(
                f"这道题适合按“识别考点 -> 套用规则 -> 结果复核”的三步来做，"
                f"核心是先把 {topic.name} 的定义和题干条件对上。"
            ),
            step_cards=step_cards,
            misconception_alerts=misconception_alerts,
            next_drills=next_drills,
            similar_questions=self.similar_questions(question.id),
        )

    def similar_questions(self, question_id: str, limit: int = 3) -> list[SimilarQuestionView]:
        question = self.get_question(question_id)
        if not question:
            return []
        candidates = [
            item
            for item in self.list_questions_by_topic(question.topic_id)
            if item.id != question.id
        ]
        scored = []
        base_tags = set(question.tags or [])
        base_tokens = self._tokenize(question.stem)
        for item in candidates:
            difficulty_gap = abs(item.difficulty - question.difficulty)
            tag_overlap = len(base_tags & set(item.tags or []))
            token_overlap = len(base_tokens & self._tokenize(item.stem))
            score = tag_overlap * 2.0 + token_overlap * 0.3 - difficulty_gap
            reason_parts = []
            if tag_overlap:
                reason_parts.append("同标签考点")
            if token_overlap:
                reason_parts.append("题干结构相近")
            if difficulty_gap <= 0.15:
                reason_parts.append("难度接近")
            scored.append(
                (
                    score,
                    SimilarQuestionView(
                        question_id=item.id,
                        topic_id=item.topic_id,
                        stem=item.stem,
                        difficulty=item.difficulty,
                        question_type=item.question_type,
                        recommendation_reason="，".join(reason_parts) or "同一知识点延伸练习",
                    ),
                )
            )
        scored.sort(key=lambda item: item[0], reverse=True)
        return [item[1] for item in scored[:limit]]

    def _tokenize(self, text: str) -> set[str]:
        return {
            token
            for token in re.split(r"[\s，。；、,.!?：:（）()\[\]\-_/]+", text.lower())
            if token
        }

    def _evaluate_answer(self, question: Question, student_answer: str, blank_answers: list[str] | None = None) -> dict:
        correct_answer = question.answer
        explanation = question.explanation
        stem = question.stem
        normalized_student = self._normalize_text(student_answer)
        normalized_correct = self._normalize_text(correct_answer)
        normalized_blank_answers = self._normalize_blank_answers(blank_answers)
        if question.question_type == QuestionType.judgment:
            return self._evaluate_judgment(question, student_answer)

        if question.question_type == QuestionType.choice:
            return self._evaluate_choice(question, student_answer)

        if question.question_type == QuestionType.steps:
            return self._evaluate_rubric(question, student_answer, strict=False)

        if question.question_type == QuestionType.solution:
            llm_result = self._evaluate_with_llm(question, student_answer)
            if llm_result:
                return llm_result
            return self._pending_teacher_review(question, student_answer, "AI 判分不可用，解答题需要教师复核。")

        if normalized_student == normalized_correct:
            return {
                "is_correct": True,
                "score": 1.0,
                "earned_points": 1.0,
                "total_points": 1.0,
                "score_label": "1/1",
                "method": "exact",
                "feedback": "答案与参考答案一致。",
                "breakdown": [
                    ScorePointResult(
                        title="答案匹配",
                        points=1.0,
                        earned_points=1.0,
                        status="正确",
                        evidence="答案与参考答案完全一致",
                    )
                ],
            }

        blank_answer_count = len(normalized_blank_answers) if normalized_blank_answers is not None else 0
        if question.blank_count > 1 or blank_answer_count > 1:
            return self._evaluate_multi_blank(question, student_answer, normalized_blank_answers)

        keyword_match = self._evaluate_single_blank_keyword(question, student_answer)
        if keyword_match:
            return keyword_match

        if self._is_subjective(stem, correct_answer):
            llm_result = self._evaluate_with_llm(question, student_answer)
            if llm_result:
                return llm_result
            return self._pending_teacher_review(question, student_answer, "AI 判分不可用，简答题需要教师复核。")

        llm_result = self._evaluate_with_llm(question, student_answer)
        if llm_result:
            return llm_result
        return self._pending_teacher_review(question, student_answer, "规则判分未能可靠命中，已转交教师复核。")

    def _evaluate_choice(self, question: Question, student_answer: str) -> dict:
        normalized_student = self._normalize_text(student_answer)
        accepted = {self._normalize_text(question.answer)}
        for option in question.options:
            if self._normalize_text(option.key) == self._normalize_text(question.answer):
                accepted.add(self._normalize_text(option.content))
            if self._normalize_text(option.content) == self._normalize_text(question.answer):
                accepted.add(self._normalize_text(option.key))
        is_correct = normalized_student in accepted
        return {
            "is_correct": is_correct,
            "score": 1.0 if is_correct else 0.0,
            "earned_points": 1.0 if is_correct else 0.0,
            "total_points": 1.0,
            "score_label": "1/1" if is_correct else "0/1",
            "method": "choice",
            "feedback": "系统按选项答案精确匹配判分。",
            "breakdown": [
                ScorePointResult(
                    title="选项判断",
                    points=1.0,
                    earned_points=1.0 if is_correct else 0.0,
                    status="正确" if is_correct else "错误",
                    evidence=f"参考答案为 {question.answer}",
                )
            ],
        }

    def _evaluate_judgment(self, question: Question, student_answer: str) -> dict:
        normalized_student = self._normalize_judgment_answer(student_answer)
        normalized_correct = self._normalize_judgment_answer(question.answer)
        if normalized_correct is None:
            normalized_correct = self._normalize_text(question.answer)
        student_value = normalized_student if normalized_student is not None else self._normalize_text(student_answer)
        is_correct = student_value == normalized_correct
        display_answer = "正确" if normalized_correct is True else ("错误" if normalized_correct is False else question.answer)
        return {
            "is_correct": is_correct,
            "score": 1.0 if is_correct else 0.0,
            "earned_points": 1.0 if is_correct else 0.0,
            "total_points": 1.0,
            "score_label": "1/1" if is_correct else "0/1",
            "method": "judgment",
            "feedback": "系统按判断题同义表达判分，支持正确/错误、对/错、true/false 等写法。",
            "breakdown": [
                ScorePointResult(
                    title="判断结果",
                    points=1.0,
                    earned_points=1.0 if is_correct else 0.0,
                    status="正确" if is_correct else "错误",
                    evidence=f"参考判断为 {display_answer}",
                )
            ],
        }

    def _normalize_judgment_answer(self, value: str) -> bool | str | None:
        raw = (value or "").strip().lower()
        normalized = self._normalize_text(raw)
        normalized = normalized.replace(".", "").replace("。", "").replace("！", "").replace("!", "")
        true_values = {"正确", "对", "是", "true", "t", "yes", "y", "√", "✓", "1", "正确的", "对的"}
        false_values = {"错误", "错", "否", "false", "f", "no", "n", "×", "✗", "x", "0", "错误的", "错的", "不正确"}
        if normalized in true_values:
            return True
        if normalized in false_values:
            return False
        if "不正确" in normalized or "错误" in normalized:
            return False
        if normalized.endswith("是正确的") or normalized.endswith("正确"):
            return True
        if normalized.endswith("是错误的") or normalized.endswith("错误"):
            return False
        return None

    def _evaluate_multi_blank(
        self,
        question: Question,
        student_answer: str,
        blank_answers: list[str] | None = None,
    ) -> dict:
        correct_parts = self._split_correct_blank_answers(question)
        student_parts = blank_answers if blank_answers is not None else self._split_student_blank_answers(
            student_answer,
            question.blank_count,
        )
        breakdown = []
        hits = 0
        total = max(len(correct_parts), question.blank_count, 1)
        for index in range(total):
            correct_part = correct_parts[index] if index < len(correct_parts) else ""
            student_part = student_parts[index] if index < len(student_parts) else ""
            accepted_values = self._accepted_blank_values(question, index, correct_part)
            matched = self._blank_answer_matches(student_part, accepted_values)
            if matched:
                hits += 1
            title = question.score_points[index].title if index < len(question.score_points) else f"第 {index + 1} 空"
            breakdown.append(
                ScorePointResult(
                    title=title,
                    points=1.0,
                    earned_points=1.0 if matched else 0.0,
                    status="命中" if matched else "未命中",
                    evidence=correct_part or "请补全对应答案",
                )
            )
        score = hits / total
        return {
            "is_correct": score >= 0.99,
            "score": score,
            "earned_points": float(hits),
            "total_points": float(total),
            "score_label": f"{hits}/{total}",
            "method": "multi_blank",
            "feedback": "系统按多空填答逐项比对，并接受关键值简写。",
            "breakdown": breakdown,
        }

    def _evaluate_single_blank_keyword(self, question: Question, student_answer: str) -> dict | None:
        accepted_values = self._accepted_blank_values(question, 0, question.answer)
        if not self._blank_answer_matches(student_answer, accepted_values):
            return None
        if not self._single_blank_core_safe(question.answer, student_answer):
            return None
        return {
            "is_correct": True,
            "score": 1.0,
            "earned_points": 1.0,
            "total_points": 1.0,
            "score_label": "1/1",
            "method": "keyword_match",
            "feedback": "答案命中了参考答案的核心表达。",
            "breakdown": [
                ScorePointResult(
                    title="答案匹配",
                    points=1.0,
                    earned_points=1.0,
                    status="命中",
                    evidence=f"核心表达：{student_answer.strip()}",
                )
            ],
        }

    def _single_blank_core_safe(self, correct_answer: str, student_answer: str) -> bool:
        normalized_correct = self._normalize_text(correct_answer)
        normalized_student = self._normalize_text(student_answer)
        if not normalized_student:
            return False
        if re.search(r"\d", normalized_correct) and not re.search(r"\d", normalized_student):
            return False
        negative_markers = ["没有", "不能", "不是", "不", "无", "非"]
        if any(marker in correct_answer for marker in negative_markers) and not any(
            marker in student_answer for marker in negative_markers
        ):
            return False
        if self._answer_has_multiple_requirements(correct_answer) and normalized_correct not in normalized_student:
            return False
        if (
            normalized_student.isascii()
            and normalized_student.isalpha()
            and re.search(r"[a-zA-Z]", normalized_correct)
            and normalized_student != normalized_correct
        ):
            return False
        return len(normalized_student) >= 2 or re.search(r"\d", normalized_student) is not None

    def _answer_has_multiple_requirements(self, answer: str) -> bool:
        if "或" in answer:
            return False
        return bool(re.search(r"[，,；;、/]|和|与|以及|分别|至少", answer))

    def _normalize_blank_answers(self, blank_answers: list[str] | None) -> list[str] | None:
        if blank_answers is None:
            return None
        return [(answer or "").strip() for answer in blank_answers]

    def _split_correct_blank_answers(self, question: Question) -> list[str]:
        parts = [part.strip() for part in re.split(r"[，,；;、]", question.answer) if part.strip()]
        if len(parts) >= question.blank_count:
            return parts
        if question.score_points:
            inferred = []
            for point in question.score_points:
                if point.keywords:
                    inferred.append(point.keywords[0])
            if len(inferred) >= question.blank_count:
                return inferred
        return parts or [question.answer]

    def _split_student_blank_answers(self, student_answer: str, expected_count: int) -> list[str]:
        answer = student_answer.strip()
        if not answer:
            return []
        if re.search(r"[，,；;、\n]", answer):
            return [part.strip() for part in re.split(r"[，,；;、\n]", answer)]
        whitespace_parts = [part.strip() for part in answer.split() if part.strip()]
        if expected_count > 1 and 1 < len(whitespace_parts) <= expected_count:
            return whitespace_parts
        return [answer]

    def _accepted_blank_values(self, question: Question, index: int, correct_part: str) -> set[str]:
        accepted = {correct_part}
        if index < len(question.score_points):
            accepted.update(question.score_points[index].keywords)
        accepted.update(self._extract_core_values(correct_part))
        return {value for value in accepted if value and value.strip()}

    def _extract_core_values(self, text: str) -> set[str]:
        values = set()
        clean = text.strip()
        if not clean:
            return values
        for part in re.split(r"或|/|｜|\|", clean):
            part = part.strip(" （）()")
            if part:
                values.add(part)
        for match in re.findall(r"(?:是|为|=|＝)\s*([^，,；;、。()（）]+)", clean):
            if match.strip():
                values.add(match.strip())
        for number in re.findall(r"(?<![\w.])-?\d+(?:\.\d+)?", clean):
            values.add(number)
        for equation in re.findall(r"[a-zA-Z]\s*[=＝]\s*[^，,；;、。()（）]+", clean):
            values.add(equation.strip())
        return values

    def _blank_answer_matches(self, student_part: str, accepted_values: set[str]) -> bool:
        normalized_student = self._normalize_text(student_part)
        if not normalized_student:
            return False
        for value in accepted_values:
            normalized_value = self._normalize_text(value)
            if not normalized_value:
                continue
            if normalized_student == normalized_value:
                return True
            if len(normalized_student) <= 8 and normalized_student in normalized_value:
                return True
            if len(normalized_value) <= 8 and normalized_value in normalized_student:
                return True
        return False

    def _evaluate_rubric(self, question: Question, student_answer: str, strict: bool) -> dict:
        score_points = question.score_points or self._infer_score_points(question.question_type, question.explanation, question.answer)
        breakdown = []
        earned_points = 0.0
        total_points = round(sum(point.points for point in score_points), 2) or 10.0
        student_tokens = self._tokenize(student_answer)
        normalized_answer = self._normalize_text(student_answer)
        for point in score_points:
            keywords = point.keywords or list(self._tokenize(point.title))
            matched_keywords = [keyword for keyword in keywords if self._normalize_text(keyword) in normalized_answer or self._normalize_text(keyword) in student_tokens]
            coverage = min(len(matched_keywords) / max(len(keywords), 1), 1.0)
            earned = point.points if coverage >= 0.99 else round(point.points * coverage, 2)
            if strict and coverage < 0.45:
                earned = 0.0
            earned_points += earned
            breakdown.append(
                ScorePointResult(
                    title=point.title,
                    points=point.points,
                    earned_points=earned,
                    status="已覆盖" if earned >= point.points * 0.8 else ("部分覆盖" if earned > 0 else "待补充"),
                    evidence="、".join(matched_keywords[:3]) if matched_keywords else "答案中尚未体现该得分点",
                )
            )
        score = round(earned_points / total_points, 4) if total_points else 0.0
        return {
            "is_correct": score >= (0.8 if strict else 0.7),
            "score": score,
            "earned_points": round(earned_points, 2),
            "total_points": total_points,
            "score_label": f"{round(earned_points, 1)}/{round(total_points, 1)}",
            "method": "rubric_steps" if question.question_type == QuestionType.steps else "rubric_solution",
            "feedback": "系统按得分点覆盖度给出学习型评分，适合看步骤完成度和订正方向。",
            "breakdown": breakdown,
        }

    def _pending_teacher_review(self, question: Question, student_answer: str, reason: str) -> dict:
        total_points = 10.0 if question.question_type in {QuestionType.solution, QuestionType.steps} else 1.0
        return {
            "is_correct": False,
            "score": 0.0,
            "earned_points": 0.0,
            "total_points": total_points,
            "score_label": "待复核",
            "method": "pending_teacher_review",
            "feedback": "这次答案已进入教师复核，不会暂时计入掌握度。",
            "breakdown": [
                ScorePointResult(
                    title="待教师复核",
                    points=total_points,
                    earned_points=0.0,
                    status="待复核",
                    evidence=reason,
                )
            ],
            "review_status": "pending_review",
            "review_reason": reason,
        }

    def _build_step_cards(self, question: Question, topic_name: str) -> list[WorkedStep]:
        if question.score_points:
            cards = []
            for point in question.score_points[:4]:
                cues = f" 关键词：{'、'.join(point.keywords[:3])}" if point.keywords else ""
                cards.append(WorkedStep(title=point.title, content=f"先完成这一得分点，再继续往下推。{cues}".strip()))
            if cards:
                return cards
        return [
            WorkedStep(
                title="先定位考点",
                content=(
                    f"这道题属于“{topic_name}”。先用题干里的关键词锁定考点："
                    f"{'、'.join(question.tags[:3]) or topic_name}"
                ),
            ),
            WorkedStep(
                title="再套用规则",
                content=(
                    question.explanation
                    or f"优先使用 {topic_name} 的核心规则，并把题干条件逐一对应到公式或定义。"
                ),
            ),
            WorkedStep(
                title="最后做复核",
                content=(
                    f"把你的结果和标准答案“{question.answer}”对照，检查单位、符号、定义是否一致。"
                ),
            ),
        ]

    def _question_from_row(self, row: QuestionBankORM) -> Question:
        question_type = self._resolve_question_type(row.question_type, row.stem, row.answer, row.score_points or [])
        knowledge_l2_id = (row.knowledge_l2_id or row.topic_id or "").strip()
        knowledge_l1_id = (row.knowledge_l1_id or "").strip()
        difficulty_level = int(row.difficulty_level or self._difficulty_float_to_level(row.difficulty))
        difficulty = float(row.difficulty if row.difficulty is not None else self._difficulty_level_to_float(difficulty_level))
        return Question(
            id=row.external_id,
            topic_id=knowledge_l2_id,
            knowledge_l1_id=knowledge_l1_id,
            knowledge_l2_id=knowledge_l2_id,
            stem=row.stem,
            difficulty_level=difficulty_level,
            difficulty=difficulty,
            answer=row.answer,
            explanation=row.explanation,
            knowledge_tiers=row.knowledge_tiers or ["基础知识点"],
            question_type=question_type,
            options=row.options or self._infer_options(question_type, row.stem),
            blank_count=row.blank_count or self._infer_blank_count(row.answer),
            score_points=row.score_points or self._infer_score_points(question_type, row.explanation, row.answer),
            tags=row.tags or [],
        )

    def _resolve_question_type(self, raw_type: str | None, stem: str, answer: str, score_points: list) -> QuestionType:
        normalized_type = self._normalize_question_type_alias(raw_type)
        if normalized_type is not None:
            return normalized_type
        if score_points:
            return QuestionType.steps
        if self._normalize_judgment_answer(answer) is not None or re.search(r"判断|正确|错误|对错|是否", stem):
            return QuestionType.judgment
        if re.search(r"[A-D][.．、]", stem):
            return QuestionType.choice
        if self._is_subjective(stem, answer):
            return QuestionType.solution
        if self._is_multi_part(answer):
            return QuestionType.blank
        return QuestionType.blank

    def _normalize_question_type_alias(self, raw_type: str | None) -> QuestionType | None:
        if not raw_type:
            return None
        normalized = raw_type.strip().lower()
        if not normalized:
            return None
        type_aliases = {
            "选择": QuestionType.choice,
            "选择题": QuestionType.choice,
            "单选": QuestionType.choice,
            "单选题": QuestionType.choice,
            "多选": QuestionType.choice,
            "多选题": QuestionType.choice,
            "single_choice": QuestionType.choice,
            "multiple_choice": QuestionType.choice,
            "choice": QuestionType.choice,
            "判断": QuestionType.judgment,
            "判断题": QuestionType.judgment,
            "true_false": QuestionType.judgment,
            "tf": QuestionType.judgment,
            "judgment": QuestionType.judgment,
            "填空": QuestionType.blank,
            "填空题": QuestionType.blank,
            "blank_question": QuestionType.blank,
            "blank": QuestionType.blank,
            "简答": QuestionType.solution,
            "简答题": QuestionType.solution,
            "解答": QuestionType.solution,
            "解答题": QuestionType.solution,
            "solution_question": QuestionType.solution,
            "solution": QuestionType.solution,
            "分步": QuestionType.steps,
            "分步题": QuestionType.steps,
            "分步计算题": QuestionType.steps,
            "step": QuestionType.steps,
            "steps": QuestionType.steps,
        }
        if normalized in type_aliases:
            return type_aliases[normalized]
        try:
            return QuestionType(normalized)
        except ValueError:
            return None

    def _infer_options(self, question_type: QuestionType, stem: str) -> list[dict]:
        if question_type != QuestionType.choice:
            return []
        matches = re.findall(r"([A-D])[.．、]\s*([^A-D]+?)(?=(?:[A-D][.．、])|$)", stem)
        return [{"key": key, "content": content.strip()} for key, content in matches]

    def _infer_blank_count(self, answer: str) -> int:
        return max(len([part for part in re.split(r"[，,；;、]", answer) if part.strip()]), 1)

    def _infer_score_points(self, question_type: QuestionType, explanation: str, answer: str) -> list[dict]:
        if question_type not in {QuestionType.solution, QuestionType.steps}:
            return []
        source = explanation or answer
        sentences = [item.strip() for item in re.split(r"[。；;]", source) if item.strip()]
        points = []
        for index, sentence in enumerate(sentences[:3], start=1):
            points.append(
                {
                    "title": f"步骤 {index}",
                    "points": round(10 / max(len(sentences[:3]), 1), 1),
                    "keywords": list(self._tokenize(sentence))[:3],
                }
            )
        return points

    def _dump_model(self, item) -> dict:
        if hasattr(item, "model_dump"):
            return item.model_dump()
        return item.dict()

    def _views(self, session, rows: list[QuestionBankORM]) -> list[QuestionBankItemView]:
        meta_map = self._build_question_meta_map(session, rows)
        return [self._view(session, row, meta_map=meta_map) for row in rows]

    def _build_question_meta_map(self, session, rows: list[QuestionBankORM]) -> dict[str, dict]:
        keys = {
            key.strip()
            for row in rows
            for key in [row.knowledge_l1_id or "", row.knowledge_l2_id or row.topic_id or "", row.topic_id or ""]
            if key and key.strip()
        }
        if not keys:
            return {}
        nodes = session.execute(
            select(KnowledgeNodeORM).where(
                KnowledgeNodeORM.is_deleted == 0,
                or_(KnowledgeNodeORM.node_key.in_(keys), KnowledgeNodeORM.topic_ref_id.in_(keys)),
            )
        ).scalars().all()
        meta_map: dict[str, dict] = {}
        for node in nodes:
            meta = {
                "name": node.name,
                "subject": node.subject or "",
                "grade_level": node.grade_level or "",
                "parent_id": node.parent_node_key or "",
                "level": node.level,
            }
            meta_map[node.node_key] = meta
            if node.topic_ref_id:
                meta_map[node.topic_ref_id] = meta
        return meta_map

    def _normalize_text(self, text: str) -> str:
        normalized = text.strip().lower()
        normalized = normalized.replace(" ", "").replace("，", ",").replace("。", "")
        normalized = normalized.replace("＝", "=").replace("：", ":")
        return normalized

    def _is_multi_part(self, answer: str) -> bool:
        return len(re.split(r"[，,；;、]", answer)) > 1

    def _is_subjective(self, stem: str, answer: str) -> bool:
        flags = ["说明", "证明", "分析", "解答", "为什么", "理由"]
        return any(flag in stem for flag in flags) or len(answer) >= 20

    def _view(self, session, row: QuestionBankORM, meta_map: dict[str, dict] | None = None) -> QuestionBankItemView:
        question = self._question_from_row(row)
        meta_map = meta_map or self._build_question_meta_map(session, [row])
        resolved_l2_id = question.knowledge_l2_id or question.topic_id
        l2_meta = meta_map.get(resolved_l2_id, {})
        resolved_l1_id = (question.knowledge_l1_id or l2_meta.get("parent_id") or "").strip()
        l1_meta = meta_map.get(resolved_l1_id, {}) if resolved_l1_id else {}

        knowledge_l2_name = l2_meta.get("name") or ""
        subject = l2_meta.get("subject") or l1_meta.get("subject") or ""
        grade_level = l2_meta.get("grade_level") or l1_meta.get("grade_level") or ""
        if not knowledge_l2_name and resolved_l2_id:
            topic = self._safe_topic(resolved_l2_id)
            knowledge_l2_name = topic.name
            subject = subject or topic.subject
            grade_level = grade_level or topic.grade_level
            resolved_l1_id = resolved_l1_id or (topic.parent_id or "")

        knowledge_l1_name = l1_meta.get("name") or ""
        if not knowledge_l1_name and resolved_l1_id:
            parent_topic = self._safe_topic(resolved_l1_id)
            knowledge_l1_name = parent_topic.name
            subject = subject or parent_topic.subject
            grade_level = grade_level or parent_topic.grade_level

        return QuestionBankItemView(
            id=row.id,
            external_id=row.external_id,
            knowledge_l1_id=resolved_l1_id,
            knowledge_l1_name=knowledge_l1_name,
            knowledge_l2_id=resolved_l2_id,
            knowledge_l2_name=knowledge_l2_name or resolved_l2_id,
            topic_id=resolved_l2_id,
            topic_name=knowledge_l2_name or resolved_l2_id,
            subject=subject,
            grade_level=grade_level,
            stem=row.stem,
            difficulty_level=question.difficulty_level,
            difficulty=question.difficulty,
            answer=row.answer,
            explanation=row.explanation,
            knowledge_tiers=question.knowledge_tiers or ["基础知识点"],
            question_type=question.question_type,
            options=question.options,
            blank_count=question.blank_count,
            score_points=question.score_points,
            tags=question.tags,
            status=row.status or "approved",
            source=row.source or "seed",
        )

    def generate_questions(self, request: QuestionGenerateRequest) -> QuestionGenerateResponse:
        topic = self._safe_topic(request.knowledge_l2_id)
        min_level = int(request.difficulty_level_min)
        max_level = int(request.difficulty_level_max)
        difficulty_min = self._difficulty_level_to_float(min_level)
        difficulty_max = self._difficulty_level_to_float(max_level)
        raw_questions = self.llm_service.generate_questions(
            topic_name=topic.name,
            subject=topic.subject,
            subtopics=topic.subtopics,
            count=request.count,
            difficulty_min=difficulty_min,
            difficulty_max=difficulty_max,
            question_type=request.question_type.value,
        )
        if not raw_questions:
            return QuestionGenerateResponse(generated_count=0, questions=[])
        if isinstance(raw_questions, dict):
            raw_questions = [raw_questions]

        saved = []
        with sql_repository.session() as session:
            knowledge_l1_id, knowledge_l2_id = self._validate_knowledge_binding(
                session,
                request.knowledge_l1_id or "",
                request.knowledge_l2_id,
            )
            for item in raw_questions:
                external_id = f"ai_{knowledge_l2_id}_{uuid.uuid4().hex[:8]}"
                options = item.get("options", [])
                score_points = item.get("score_points", [])
                blank_count = item.get("blank_count", 1)
                if options:
                    blank_count = 1
                generated_level = self._coerce_difficulty_level(
                    item.get("difficulty_level"),
                    item.get("difficulty"),
                    fallback=min(max_level, max(min_level, 3)),
                )
                row = QuestionBankORM(
                    external_id=external_id,
                    knowledge_l1_id=knowledge_l1_id,
                    knowledge_l2_id=knowledge_l2_id,
                    topic_id=knowledge_l2_id,
                    stem=item.get("stem", ""),
                    difficulty_level=generated_level,
                    difficulty=self._difficulty_level_to_float(generated_level),
                    answer=item.get("answer", ""),
                    explanation=item.get("explanation", "") if request.include_explanation else "",
                    knowledge_tiers=self._normalize_knowledge_tiers(item.get("knowledge_tiers")),
                    question_type=request.question_type.value,
                    options=options,
                    blank_count=blank_count,
                    score_points=score_points,
                    tags=item.get("tags", []),
                    status="pending",
                    source="ai_generated",
                )
                session.add(row)
                session.flush()
                saved.append(self._view(session, row))

        return QuestionGenerateResponse(generated_count=len(saved), questions=saved)

    def review_questions(self, request: QuestionReviewRequest) -> QuestionReviewResponse:
        new_status = "approved" if request.action == "approve" else "rejected"
        reviewed = []
        with sql_repository.session() as session:
            for qid in request.question_ids:
                row = session.execute(
                    select(QuestionBankORM).where(QuestionBankORM.id == qid)
                ).scalars().first()
                if row and row.status == "pending":
                    row.status = new_status
                    session.flush()
                    reviewed.append(self._view(session, row))
        return QuestionReviewResponse(reviewed_count=len(reviewed), questions=reviewed)

    def list_pending_questions(self) -> list[QuestionBankItemView]:
        with sql_repository.session() as session:
            rows = session.execute(
                select(QuestionBankORM).where(QuestionBankORM.status == "pending").order_by(QuestionBankORM.created_at.desc())
            ).scalars().all()
            return self._views(session, rows)

    def import_csv(self, csv_content: str, teacher_user_id: int | None = None) -> CsvImportResponse:
        cn_to_en = {
            "题目": "stem", "题干": "stem",
            "答案": "answer",
            "一级知识点ID": "knowledge_l1_id", "一级知识点": "knowledge_l1_id",
            "一级知识点名称": "knowledge_l1_name",
            "二级知识点ID": "knowledge_l2_id", "二级知识点": "knowledge_l2_id",
            "二级知识点名称": "knowledge_l2_name",
            "知识点ID": "knowledge_l2_id", "知识点": "knowledge_l2_id",
            "知识点名称": "knowledge_l2_name",
            "难度级别": "difficulty_level", "难度等级": "difficulty_level",
            "难度": "difficulty_level",
            "解析": "explanation",
            "题型": "question_type",
            "选项": "options",
            "空数": "blank_count", "填空数": "blank_count",
            "得分点": "score_points",
            "知识点层级标签": "knowledge_tiers",
            "标签": "tags",
            "编号": "id",
            "knowledge_l1_name": "knowledge_l1_name",
            "knowledge_l2_name": "knowledge_l2_name",
        }
        reader = csv.DictReader(io.StringIO(csv_content))
        if reader.fieldnames:
            mapped = []
            for f in reader.fieldnames:
                field_name = f.strip().lstrip("\ufeff")
                mapped.append(cn_to_en.get(field_name, field_name))
            reader = csv.DictReader(io.StringIO(csv_content), fieldnames=mapped)
            next(reader)
        imported = []
        skipped = 0
        with sql_repository.session() as session:
            csv_context = self._build_csv_topic_context(session, teacher_user_id)
            for row_data in reader:
                stem = row_data.get("stem", "").strip()
                answer = row_data.get("answer", "").strip()
                knowledge_l1_id = row_data.get("knowledge_l1_id", "").strip()
                knowledge_l1_name = row_data.get("knowledge_l1_name", "").strip()
                knowledge_l2_id = row_data.get("knowledge_l2_id", "").strip()
                knowledge_l2_name = row_data.get("knowledge_l2_name", "").strip()
                if not stem or not answer or (not knowledge_l2_id and not knowledge_l2_name):
                    skipped += 1
                    continue
                external_id = row_data.get("id", f"csv_{uuid.uuid4().hex[:8]}").strip()
                existing = session.execute(
                    select(QuestionBankORM).where(QuestionBankORM.external_id == external_id)
                ).scalars().first()
                if existing:
                    skipped += 1
                    continue
                try:
                    difficulty_level = self._coerce_difficulty_level(row_data.get("difficulty_level"), None)
                except ValueError:
                    skipped += 1
                    continue
                explanation = row_data.get("explanation", "")
                raw_question_type = row_data.get("question_type", "blank").strip()
                question_type = self._resolve_question_type(raw_question_type, stem, answer, []).value
                tags_str = row_data.get("tags", "")
                tags = [t.strip() for t in tags_str.split(",") if t.strip()] if tags_str else []
                tiers_raw = row_data.get("knowledge_tiers", "")
                tier_parts = [t.strip() for t in re.split(r"[,，|、;/；]", tiers_raw) if t.strip()] if tiers_raw else []
                knowledge_tiers = self._normalize_knowledge_tiers(tier_parts)
                options = self._parse_csv_options(row_data.get("options", ""))
                if options:
                    question_type = QuestionType.choice.value
                if question_type == QuestionType.choice.value and not options:
                    options = self._infer_options(QuestionType.choice, stem)
                score_points = self._parse_csv_score_points(row_data.get("score_points", ""))
                blank_count = self._parse_csv_blank_count(row_data.get("blank_count", ""), answer, score_points)
                if question_type in {QuestionType.choice.value, QuestionType.judgment.value, QuestionType.solution.value}:
                    blank_count = 1
                try:
                    resolved_l1_id, resolved_l2_id = self._resolve_csv_knowledge_binding(
                        session,
                        knowledge_l1_id=knowledge_l1_id,
                        knowledge_l1_name=knowledge_l1_name,
                        knowledge_l2_id=knowledge_l2_id,
                        knowledge_l2_name=knowledge_l2_name,
                        context=csv_context,
                    )
                except ValueError:
                    skipped += 1
                    continue
                row = QuestionBankORM(
                    external_id=external_id,
                    knowledge_l1_id=resolved_l1_id,
                    knowledge_l2_id=resolved_l2_id,
                    topic_id=resolved_l2_id,
                    stem=stem,
                    difficulty_level=difficulty_level,
                    difficulty=self._difficulty_level_to_float(difficulty_level),
                    answer=answer,
                    explanation=explanation,
                    knowledge_tiers=knowledge_tiers,
                    question_type=question_type,
                    options=options,
                    blank_count=blank_count,
                    score_points=score_points,
                    tags=tags,
                    status="pending",
                    source="csv_import",
                )
                session.add(row)
                session.flush()
                imported.append(self._view(session, row))
        return CsvImportResponse(imported_count=len(imported), skipped_count=skipped, questions=imported)

    def build_excel_template(
        self,
        teacher_user_id: int,
        grade_level: str,
        subject: str,
        row_count: int = 12,
    ) -> tuple[str, bytes]:
        normalized_grade = (grade_level or "").strip()
        normalized_subject = (subject or "").strip()
        if not normalized_grade or not normalized_subject:
            raise ValueError("年级和学科不能为空")

        with sql_repository.session() as session:
            context = self._build_topic_context(
                session,
                teacher_user_id=teacher_user_id,
                grade_level=normalized_grade,
                subject=normalized_subject,
                strict_grade_scope=True,
            )
            topic_names = [item["knowledge_l2_name"] for item in context["topics"]]
            if not topic_names:
                raise ValueError("当前年级学科下暂无可用二级知识点")
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "题目录入"
        meta = workbook.create_sheet("_meta")
        meta["A1"] = "ZHUYU_EDU_QUESTION_IMPORT_TEMPLATE"
        meta["A2"] = normalized_grade
        meta["A3"] = normalized_subject
        meta["A4"] = "一级知识点固定为当前学科，二级知识点请从下拉中选择"
        meta.sheet_state = "hidden"

        for index, header in enumerate(self._excel_template_headers, start=1):
            sheet.cell(row=1, column=index, value=header)
            sheet.column_dimensions[self._excel_column_name(index)].width = 18 if index != 1 else 36
        sheet.freeze_panes = "A2"

        for row_index, name in enumerate(topic_names, start=1):
            meta.cell(row=row_index, column=2, value=name)

        options_range = f"'_meta'!$B$1:$B${max(len(topic_names), 1)}"
        validation = DataValidation(type="list", formula1=options_range, allow_blank=False)
        validation.prompt = "请选择当前年级学科下的二级知识点"
        validation.error = "请从下拉选项中选择二级知识点"
        sheet.add_data_validation(validation)
        validation.add(f"C2:C{row_count + 1}")

        for row_index in range(2, row_count + 2):
            sheet.cell(row=row_index, column=4, value="")
            sheet.cell(row=row_index, column=6, value=3)
            sheet.cell(row=row_index, column=8, value=1)

        buffer = io.BytesIO()
        workbook.save(buffer)
        filename = f"题目导入模板_{normalized_grade}_{normalized_subject}.xlsx"
        return filename, buffer.getvalue()

    def import_excel(
        self,
        content: bytes,
        teacher_user_id: int,
        grade_level: str,
        subject: str,
    ) -> CsvImportResponse:
        normalized_grade = (grade_level or "").strip()
        normalized_subject = (subject or "").strip()
        if not normalized_grade or not normalized_subject:
            raise ValueError("年级和学科不能为空")

        from openpyxl import load_workbook

        try:
            workbook = load_workbook(io.BytesIO(content))
        except Exception as exc:
            raise ValueError("Excel 文件无法解析，请使用系统模板") from exc

        self._validate_excel_template(workbook, normalized_grade, normalized_subject)
        if "题目录入" not in workbook.sheetnames:
            raise ValueError("Excel 模板中缺少题目录入工作表")
        sheet = workbook["题目录入"]
        headers = [self._excel_cell_text(sheet.cell(row=1, column=index).value) for index in range(1, len(self._excel_template_headers) + 1)]
        if headers != self._excel_template_headers:
            raise ValueError("Excel 模板表头不匹配，请重新下载模板")

        imported = []
        skipped = 0
        failed_rows: list[ImportFailureItem] = []
        with sql_repository.session() as session:
            context = self._build_topic_context(
                session,
                teacher_user_id=teacher_user_id,
                grade_level=normalized_grade,
                subject=normalized_subject,
                strict_grade_scope=True,
            )
            if not context["topics"]:
                raise ValueError("当前年级学科下暂无可用二级知识点")

            for row_index in range(2, sheet.max_row + 1):
                values = [
                    self._excel_cell_text(sheet.cell(row=row_index, column=column_index).value)
                    for column_index in range(1, len(self._excel_template_headers) + 1)
                ]
                stem, answer, knowledge_l2_name, raw_question_type, explanation, raw_difficulty_level, options_raw, raw_blank_count, raw_score_points, tags_raw = values[:10]
                meaningful_values = [stem, answer, knowledge_l2_name, raw_question_type, explanation, options_raw, raw_score_points, tags_raw]
                if not any(meaningful_values):
                    continue
                stem_preview = self._stem_preview(stem)
                if not stem:
                    skipped += 1
                    failed_rows.append(ImportFailureItem(row_number=row_index, reason="题目为空", stem_preview=stem_preview))
                    continue
                if not answer:
                    skipped += 1
                    failed_rows.append(ImportFailureItem(row_number=row_index, reason="答案为空", stem_preview=stem_preview))
                    continue
                if not knowledge_l2_name:
                    skipped += 1
                    failed_rows.append(ImportFailureItem(row_number=row_index, reason="二级知识点为空，请从下拉框选择", stem_preview=stem_preview))
                    continue
                if raw_question_type and self._normalize_question_type_alias(raw_question_type) is None:
                    skipped += 1
                    failed_rows.append(
                        ImportFailureItem(
                            row_number=row_index,
                            reason="题型无法识别，请填写选择题、判断题、填空题、解答题或分步计算题",
                            stem_preview=stem_preview,
                        )
                    )
                    continue
                try:
                    difficulty_level = self._coerce_difficulty_level(raw_difficulty_level, None)
                except ValueError:
                    skipped += 1
                    failed_rows.append(
                        ImportFailureItem(
                            row_number=row_index,
                            reason="难度级别必须是 1-5 的数字",
                            stem_preview=stem_preview,
                        )
                    )
                    continue

                options = self._parse_csv_options(options_raw)
                score_points = self._parse_csv_score_points(raw_score_points)
                question_type = self._resolve_question_type(raw_question_type, stem, answer, score_points).value
                if options:
                    question_type = QuestionType.choice.value
                if question_type == QuestionType.choice.value and not options:
                    options = self._infer_options(QuestionType.choice, stem)
                if question_type == QuestionType.choice.value and not options:
                    skipped += 1
                    failed_rows.append(
                        ImportFailureItem(
                            row_number=row_index,
                            reason="选择题缺少有效选项，请按 A:选项一|B:选项二 的格式填写",
                            stem_preview=stem_preview,
                        )
                    )
                    continue
                if options_raw and question_type != QuestionType.choice.value and not options:
                    skipped += 1
                    failed_rows.append(
                        ImportFailureItem(
                            row_number=row_index,
                            reason="选项格式无法识别，请按 A:选项一|B:选项二 的格式填写",
                            stem_preview=stem_preview,
                        )
                    )
                    continue
                blank_count = self._parse_csv_blank_count(raw_blank_count, answer, score_points)
                if question_type in {QuestionType.choice.value, QuestionType.judgment.value, QuestionType.solution.value}:
                    blank_count = 1
                tags = [item.strip() for item in re.split(r"[,，]", tags_raw or "") if item.strip()]
                try:
                    _, resolved_l2_id = self._resolve_csv_knowledge_binding(
                        session,
                        knowledge_l1_id="",
                        knowledge_l1_name="",
                        knowledge_l2_id="",
                        knowledge_l2_name=knowledge_l2_name,
                        context=context,
                    )
                    resolved_l1_id, resolved_l2_id = self._validate_knowledge_binding(session, "", resolved_l2_id)
                except ValueError as exc:
                    skipped += 1
                    reason = self._friendly_import_reason(exc)
                    failed_rows.append(
                        ImportFailureItem(
                            row_number=row_index,
                            reason=reason,
                            stem_preview=stem_preview,
                        )
                    )
                    continue

                row_model = QuestionBankORM(
                    external_id=f"excel_{uuid.uuid4().hex[:8]}",
                    knowledge_l1_id=resolved_l1_id,
                    knowledge_l2_id=resolved_l2_id,
                    topic_id=resolved_l2_id,
                    stem=stem,
                    difficulty_level=difficulty_level,
                    difficulty=self._difficulty_level_to_float(difficulty_level),
                    answer=answer,
                    explanation=explanation,
                    knowledge_tiers=["基础知识点"],
                    question_type=question_type,
                    options=options,
                    blank_count=blank_count,
                    score_points=score_points,
                    tags=tags,
                    status="pending",
                    source="excel_import",
                )
                session.add(row_model)
                session.flush()
                imported.append(self._view(session, row_model))
        return CsvImportResponse(
            imported_count=len(imported),
            skipped_count=skipped,
            questions=imported,
            failed_rows=failed_rows,
        )

    def _build_topic_context(
        self,
        session,
        teacher_user_id: int | None = None,
        grade_level: str | None = None,
        subject: str | None = None,
        strict_grade_scope: bool = False,
    ) -> dict:
        school_id = None
        preferred_grades: list[str] = []
        if teacher_user_id is not None:
            teacher = session.execute(select(UserORM).where(UserORM.id == teacher_user_id)).scalars().first()
            if teacher:
                school_id = teacher.school_id
            preferred_grades = [
                grade.strip()
                for grade in session.execute(
                    select(ClassroomORM.grade_level).where(ClassroomORM.teacher_user_id == teacher_user_id)
                ).scalars().all()
                if grade and grade.strip()
            ]

        stmt = select(KnowledgeNodeORM).where(KnowledgeNodeORM.is_deleted == 0, KnowledgeNodeORM.level >= 2)
        if school_id is not None:
            stmt = stmt.where(KnowledgeNodeORM.school_id == school_id)
        normalized_grade = (grade_level or "").strip()
        normalized_subject = (subject or "").strip()
        if strict_grade_scope and normalized_grade and preferred_grades and normalized_grade not in preferred_grades:
            raise ValueError("当前年级不在老师所教班级范围内")
        if normalized_grade:
            stmt = stmt.where(KnowledgeNodeORM.grade_level == normalized_grade)
        if normalized_subject:
            stmt = stmt.where(KnowledgeNodeORM.subject == normalized_subject)
        nodes = session.execute(stmt.order_by(KnowledgeNodeORM.sort_order.asc(), KnowledgeNodeORM.node_key.asc())).scalars().all()

        parents = {}
        parent_keys = {node.parent_node_key for node in nodes if node.parent_node_key}
        if parent_keys:
            parent_rows = session.execute(
                select(KnowledgeNodeORM).where(
                    KnowledgeNodeORM.is_deleted == 0,
                    KnowledgeNodeORM.node_key.in_(parent_keys),
                )
            ).scalars().all()
            parents = {row.node_key: row for row in parent_rows}

        by_name = defaultdict(list)
        topics = []
        for node in nodes:
            parent = parents.get(node.parent_node_key or "")
            item = {
                "knowledge_l2_id": node.node_key,
                "knowledge_l2_name": node.name.strip(),
                "knowledge_l1_id": node.parent_node_key or "",
                "knowledge_l1_name": parent.name.strip() if parent else "",
                "grade_level": (node.grade_level or "").strip(),
                "subject": (node.subject or "").strip(),
            }
            by_name[node.name.strip()].append(item)
            topics.append(item)
        return {
            "by_name": by_name,
            "preferred_grades": preferred_grades,
            "topics": topics,
        }

    def _build_csv_topic_context(self, session, teacher_user_id: int | None = None) -> dict:
        return self._build_topic_context(session, teacher_user_id=teacher_user_id)

    def _resolve_csv_knowledge_binding(
        self,
        session,
        *,
        knowledge_l1_id: str,
        knowledge_l1_name: str,
        knowledge_l2_id: str,
        knowledge_l2_name: str,
        context: dict,
    ) -> tuple[str, str]:
        normalized_l1_id = (knowledge_l1_id or "").strip()
        normalized_l1_name = (knowledge_l1_name or "").strip()
        normalized_l2_id = (knowledge_l2_id or "").strip()
        normalized_l2_name = (knowledge_l2_name or "").strip()

        if normalized_l2_id:
            return self._validate_knowledge_binding(session, normalized_l1_id, normalized_l2_id)

        candidates = list(context.get("by_name", {}).get(normalized_l2_name, []))
        if normalized_l1_id:
            candidates = [item for item in candidates if item["knowledge_l1_id"] == normalized_l1_id]
        if normalized_l1_name:
            candidates = [item for item in candidates if item["knowledge_l1_name"] == normalized_l1_name]
        if not candidates:
            raise ValueError("knowledge topic not found")

        preferred_grades = context.get("preferred_grades", [])
        if preferred_grades:
            preferred = [item for item in candidates if item["grade_level"] in preferred_grades]
            if preferred:
                candidates = preferred

        unique_candidates = []
        seen = set()
        for item in candidates:
            key = (item["knowledge_l1_id"], item["knowledge_l2_id"])
            if key not in seen:
                seen.add(key)
                unique_candidates.append(item)
        if len(unique_candidates) != 1:
            raise ValueError("knowledge topic ambiguous")

        candidate = unique_candidates[0]
        resolved_l1_id = normalized_l1_id or candidate["knowledge_l1_id"]
        return self._validate_knowledge_binding(session, resolved_l1_id, candidate["knowledge_l2_id"])

    def _validate_excel_template(self, workbook, grade_level: str, subject: str) -> None:
        if "_meta" not in workbook.sheetnames:
            raise ValueError("请使用系统下载的 Excel 模板")
        meta = workbook["_meta"]
        marker = self._excel_cell_text(meta["A1"].value)
        if marker != "ZHUYU_EDU_QUESTION_IMPORT_TEMPLATE":
            raise ValueError("请使用系统下载的 Excel 模板")
        template_grade = self._excel_cell_text(meta["A2"].value)
        template_subject = self._excel_cell_text(meta["A3"].value)
        if template_grade != grade_level or template_subject != subject:
            raise ValueError("Excel 模板的年级或学科与当前选择不一致，请重新下载模板")

    def _excel_cell_text(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _excel_column_name(self, index: int) -> str:
        chars = []
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            chars.append(chr(65 + remainder))
        return "".join(reversed(chars))

    def _topic_exists(self, topic_id: str) -> bool:
        if self.repository.has_topic(topic_id):
            return True
        with sql_repository.session() as session:
            row = session.execute(
                select(KnowledgeNodeORM.id).where(
                    KnowledgeNodeORM.is_deleted == 0,
                    KnowledgeNodeORM.level >= 2,
                    (KnowledgeNodeORM.node_key == topic_id) | (KnowledgeNodeORM.topic_ref_id == topic_id),
                )
            ).scalars().first()
            return bool(row)

    def _safe_topic(self, topic_id: str) -> Topic:
        try:
            return self.repository.get_topic(topic_id)
        except KeyError:
            with sql_repository.session() as session:
                node = session.execute(
                    select(KnowledgeNodeORM).where(
                        KnowledgeNodeORM.is_deleted == 0,
                        (KnowledgeNodeORM.node_key == topic_id) | (KnowledgeNodeORM.topic_ref_id == topic_id),
                    )
                ).scalars().first()
                if node:
                    return Topic(
                        id=topic_id,
                        name=node.name,
                        subject=node.subject or "",
                        parent_id=node.parent_node_key,
                        level=2 if node.level >= 2 else 1,
                        grade_level=node.grade_level or "",
                        term="全年",
                        sort_order=node.sort_order,
                        prerequisites=[],
                        subtopics=[],
                        difficulty=0.5,
                        learning_objectives=[f"掌握{node.name}"],
                        common_mistakes=[],
                        tutoring_tips=[],
                    )
            return Topic(
                id=topic_id,
                name=topic_id,
                subject="",
                parent_id=None,
                level=2,
                grade_level="",
                term="全年",
                sort_order=0,
                prerequisites=[],
                subtopics=[],
                difficulty=0.5,
                learning_objectives=[f"掌握{topic_id}"],
                common_mistakes=[],
                tutoring_tips=[],
            )

    def _parse_csv_options(self, raw: str) -> list[dict]:
        raw = (raw or "").strip()
        if not raw:
            return []
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [
                    {"key": str(item.get("key", "")).strip(), "content": str(item.get("content", "")).strip()}
                    for item in parsed
                    if isinstance(item, dict) and item.get("key") and item.get("content")
                ]
        except json.JSONDecodeError:
            pass
        options = []
        for part in re.split(r"[|｜\n\r]+", raw):
            part = part.strip()
            if not part:
                continue
            match = re.match(r"^([A-Da-d])\s*[:：.．、\)]\s*(.+)$", part)
            if match:
                options.append({"key": match.group(1).upper(), "content": match.group(2).strip()})
        return options

    def _parse_csv_blank_count(self, raw: str, answer: str, score_points: list[dict]) -> int:
        raw = (raw or "").strip()
        if raw:
            try:
                return max(int(float(raw)), 1)
            except ValueError:
                pass
        if score_points:
            return max(len(score_points), 1)
        return self._infer_blank_count(answer)

    def _parse_csv_score_points(self, raw: str) -> list[dict]:
        raw = (raw or "").strip()
        if not raw:
            return []
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [
                    {
                        "title": str(item.get("title", f"得分点 {index + 1}")).strip(),
                        "points": float(item.get("points", 1.0)),
                        "keywords": [str(keyword).strip() for keyword in item.get("keywords", []) if str(keyword).strip()],
                    }
                    for index, item in enumerate(parsed)
                    if isinstance(item, dict)
                ]
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
        points = []
        for index, part in enumerate(re.split(r"[|｜\n\r]+", raw), start=1):
            part = part.strip()
            if not part:
                continue
            pieces = [piece.strip() for piece in re.split(r"[:：]", part) if piece.strip()]
            if not pieces:
                continue
            title = pieces[0]
            point_value = 1.0
            keyword_source = pieces[-1] if len(pieces) > 1 else pieces[0]
            if len(pieces) >= 3:
                try:
                    point_value = float(pieces[1])
                except ValueError:
                    point_value = 1.0
            keywords = [item.strip() for item in re.split(r"[,，/、;；]", keyword_source) if item.strip()]
            points.append({"title": title or f"得分点 {index}", "points": point_value, "keywords": keywords})
        return points

    def _coerce_difficulty_level(self, raw_level, raw_difficulty=None, fallback: int = 3) -> int:
        if raw_level is None or str(raw_level).strip() == "":
            if raw_difficulty is not None and str(raw_difficulty).strip() != "":
                return self._difficulty_float_to_level(float(raw_difficulty))
            return max(1, min(5, int(fallback)))
        value = str(raw_level).strip()
        try:
            numeric = float(value)
        except ValueError as exc:
            raise ValueError("difficulty_level must be numeric") from exc
        if 1 <= numeric <= 5 and float(int(numeric)) == numeric:
            return int(numeric)
        if 0.0 <= numeric <= 1.0:
            return self._difficulty_float_to_level(numeric)
        rounded = int(round(numeric))
        if 1 <= rounded <= 5:
            return rounded
        raise ValueError("difficulty_level out of range")

    def _difficulty_level_to_float(self, difficulty_level: int) -> float:
        mapping = {1: 0.1, 2: 0.3, 3: 0.5, 4: 0.7, 5: 0.9}
        return mapping.get(int(difficulty_level), 0.5)

    def _difficulty_float_to_level(self, difficulty: float | None) -> int:
        if difficulty is None:
            return 3
        value = float(difficulty)
        if value < 0.2:
            return 1
        if value < 0.4:
            return 2
        if value < 0.6:
            return 3
        if value < 0.8:
            return 4
        return 5

    def _normalize_knowledge_tiers(self, tiers) -> list[str]:
        if isinstance(tiers, str):
            candidates = [item.strip() for item in re.split(r"[,，|、;/；]", tiers) if item.strip()]
        else:
            candidates = [str(item).strip() for item in (tiers or []) if str(item).strip()]
        normalized: list[str] = []
        for tier in candidates:
            if tier in self._knowledge_tier_set and tier not in normalized:
                normalized.append(tier)
        if not normalized:
            return ["基础知识点"]
        return normalized

    def _friendly_import_reason(self, error: Exception) -> str:
        raw = str(error).strip()
        if not raw:
            return "导入失败，请检查该行内容"
        lowered = raw.lower()
        if "knowledge topic not found" in lowered:
            return "二级知识点不存在，请确认使用模板下拉框中的知识点"
        if "knowledge topic ambiguous" in lowered:
            return "二级知识点名称重复，当前无法唯一匹配，请联系管理员检查知识点配置"
        return raw

    def _stem_preview(self, stem: str, limit: int = 24) -> str:
        normalized = re.sub(r"\s+", " ", (stem or "").strip())
        if not normalized:
            return ""
        if len(normalized) <= limit:
            return normalized
        return normalized[:limit] + "..."

    def _validate_knowledge_binding(self, session, knowledge_l1_id: str, knowledge_l2_id: str) -> tuple[str, str]:
        node_l2 = session.execute(
            select(KnowledgeNodeORM).where(
                KnowledgeNodeORM.is_deleted == 0,
                KnowledgeNodeORM.level >= 2,
                (KnowledgeNodeORM.node_key == knowledge_l2_id) | (KnowledgeNodeORM.topic_ref_id == knowledge_l2_id),
            )
        ).scalars().first()
        if not node_l2:
            raise ValueError("二级知识点不存在")
        resolved_l2 = node_l2.node_key
        expected_l1 = node_l2.parent_node_key or ""
        resolved_l1 = (knowledge_l1_id or expected_l1).strip()
        if not resolved_l1:
            raise ValueError("二级知识点缺少一级父节点")
        node_l1 = session.execute(
            select(KnowledgeNodeORM).where(
                KnowledgeNodeORM.is_deleted == 0,
                KnowledgeNodeORM.level == 1,
                KnowledgeNodeORM.node_key == resolved_l1,
            )
        ).scalars().first()
        if not node_l1:
            raise ValueError("一级知识点不存在")
        if expected_l1 and node_l1.node_key != expected_l1:
            raise ValueError("二级知识点不属于该一级知识点")
        return node_l1.node_key, resolved_l2

    def _evaluate_with_llm(self, question: Question, student_answer: str) -> dict:
        if not self.llm_service:
            return None
        result = self.llm_service.grade_answer(
            question_stem=question.stem,
            correct_answer=question.answer,
            student_answer=student_answer,
            explanation=question.explanation,
            allow_offline=False,
        )
        if not result:
            return None
        score = float(result.get("score", 0.0))
        is_correct = bool(result.get("is_correct", score >= 0.8))
        feedback = result.get("feedback", "")
        raw_breakdown = result.get("breakdown", [])
        breakdown = []
        for item in raw_breakdown:
            breakdown.append(
                ScorePointResult(
                    title=item.get("title", "评分项"),
                    points=float(item.get("points", 1.0)),
                    earned_points=float(item.get("earned_points", 0.0)),
                    status=item.get("status", ""),
                    evidence=item.get("evidence", ""),
                )
            )
        total_points = sum(item.points for item in breakdown) if breakdown else 1.0
        earned_points = sum(item.earned_points for item in breakdown) if breakdown else score
        return {
            "is_correct": is_correct,
            "score": score,
            "earned_points": round(earned_points, 2),
            "total_points": round(total_points, 2),
            "score_label": f"{round(earned_points, 1)}/{round(total_points, 1)}" if breakdown else f"{round(score * 100)}%",
            "method": "llm_semantic",
            "feedback": feedback or "AI 语义判分完成。",
            "breakdown": breakdown,
        }

    @staticmethod
    def _normalize_text_static(text: str) -> str:
        normalized = text.strip().lower()
        normalized = normalized.replace(" ", "").replace("，", ",").replace("。", "")
        normalized = normalized.replace("＝", "=").replace("：", ":")
        return normalized
